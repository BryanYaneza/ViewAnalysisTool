from openpyxl import Workbook
import re
import teradatasql
from datetime import date


# Load Target Table
with open("TargetView.txt") as f:
    content = f.readlines()
views = [x.strip().upper() for x in content]

# Presets
listReference = ["FROM", "INNER JOIN", "OUTER JOIN", "RIGHT OUTER JOIN", "LEFT OUTER JOIN", "CROSS JOIN"]
lockingReference = ["LOCKING TABLE", "LOCKING ROW", "LOCK TABLE", "LOCK ROW"]
list1 = []
result = []
analyzedResult = []
isTerminate = False

# Connect To Teradata
def connectToTeradata(host, username, password):
    global con
    terminate = False
    while not terminate:
        try:
            con = teradatasql.connect(
                '{"host": "' + host + '","user":"' + username + '","password":"' + password + '"}')
            terminate = True
            print("Connection Successful\n")
        except teradatasql.Error:
            print("Connection Failed\n")
            exit()
    return con

# Extract view
def extractView(consCredential, view):
    with consCredential.cursor() as cur:
        cur.execute("SHOW VIEW {0}".format(view))
        return """{}""".format("".join([row for row in cur.fetchall()][0]).replace("\r", "\n").upper())

# Drill Down on the view
def drillDown(viewDef):
    list1 = []
    for ref in listReference:
        match = re.findall('(?<=%s).*$' % ref, viewDef, re.MULTILINE)
        for x in match:
            for result in re.split(r'\s', x):
                if result == "ON":
                    break
                elif len(result.split(".")) > 1:
                    list1.append(result)
    return list1


# Test Drill Down Function
def drillDownLocal(viewDef):
    list1 = []
    for ref in listReference:
        match = re.findall('(?<=%s).*$' % ref, viewDef, re.MULTILINE)
        for x in match:
            for result in re.split(r'\s', x):
                if len(result.split(".")) > 1:
                    list1.append(result)
    return list1


# Analyze the view
def analyzeView(conn, views):
    innerList = []
    extractedViewDictionary = {}
    for view in views:
        filtView = view.replace(';', '')
        try:
            viewStructure = extractView(conn, filtView)
            for ref in lockingReference:
                innerList.append(re.findall(ref, viewStructure, re.MULTILINE))
            # Underlying view has Locking Table
            if innerList[0] or innerList[2]:
                extractedViewDictionary[filtView] = "Underlying view has Locking Table"
            # Underlying view has Locking Row
            elif innerList[1] or innerList[3]:
                extractedViewDictionary[filtView] = "Underlying view has Locking Row"
            # No Locking Statement
            else:
                extractedViewDictionary[filtView] = "No Locking Statement"
        except:
            extractedViewDictionary[filtView] = "No Locking Statement"

        innerList = []
    return extractedViewDictionary


def writeToExcel(analyzedViews, parentView):
    # print(analyzedViews)
    wb = Workbook()
    sheet = wb.active
    cnt = 2
    cnt1 = 0
    sheet.cell(row=1, column=1).value = "Parent View"
    sheet.cell(row=1, column=2).value = "Underlying Value"
    sheet.cell(row=1, column=3).value = "Analysis"
    for view in parentView:
        for underlyingView in analyzedViews[cnt1]:
            sheet.cell(row=cnt, column=1).value = view
            sheet.cell(row=cnt, column=2).value = underlyingView
            sheet.cell(row=cnt, column=3).value = analyzedViews[cnt1][underlyingView]
            cnt += 1
        cnt1 += 1

    filename = "./Results/ViewAnalysisResults-{0}.xlsx".format(date.today().strftime("%m%d%y"))
    wb.save(filename=filename)

# Main Function
def main():
    global extractedViews
    result = []
    conn = connectToTeradata(HOST, USER, PASSWORD)
    print("Running Analysis \n")
    for view in views:
        try:
            print("Analyzing view: {0}".format(view))
            extractedView = extractView(conn, view)
            # print(extractedView)
            extractedViews = drillDown(extractedView)
            print(extractedViews)
            result.append(analyzeView(conn, extractedViews))
        except teradatasql.Error:
            # print("\nError Raised.")
            result.append({extractedViews[0]:  "No Locking Statement"})

    print(result)
    # writeToExcel(result,views)
    print("\nAnalysis Done.")


print("Welcome")
main()
